"""
==========================================================
HOMS Receipt Engine
==========================================================
입고 Import Engine
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

from core.database_engine import DatabaseEngine

from datetime import datetime

# ==========================================================
# 품목 매핑
# ==========================================================

ITEM_MAP = {

    "북채": "북채",

    "통날개": "날개",
    "날개": "날개",

    "한마리 원육(900g)": "한마리",
    "한마리": "한마리",

    "허니(콤보)": "허니콤보",
    "허니콤보": "허니콤보",

    "순살안심정육(500g)": "허니순살",
    "허니순살": "허니순살",

    "순살정육(700g)": "정육순살",
    "정육순살": "정육순살",

    "윙봉(염지)_태국산(18pcs)": "태국산윙봉",
    "윙봉(염지)_태국산(20pcs)": "태국산윙봉",
    "태국산윙봉": "태국산윙봉",

    "소이(가슴)": "소이가슴",
    "소이가슴": "소이가슴",

    "소이(정육)": "소이정육",
    "소이정육": "소이정육",
}


class ReceiptEngine:

    def __init__(self):

        self.db = DatabaseEngine()

    # ------------------------------------------------------
    # 엑셀 읽기
    # ------------------------------------------------------

    def import_excel(
        self,
        file_path,
    ):

        wb = load_workbook(
            file_path,
            data_only=True,
        )

        ws = wb.worksheets[0]

        headers = {}

        header_row = None

        for r in range(1, 11):

            values = []

            for c in range(
                1,
                ws.max_column + 1,
            ):

                values.append(
                    ws.cell(
                        row=r,
                        column=c,
                    ).value
                )

            if "품목명" in values:

                header_row = r

                for col, value in enumerate(
                    values,
                    start=1,
                ):

                    if value is not None:

                        headers[
                            str(value).strip()
                        ] = col

                break

        if header_row is None:

            raise ValueError(
                "품목명 컬럼을 찾을 수 없습니다."
            )

        item_col = headers["품목명"]
        qty_col = headers["확정수량"]

        result = {}

        for row in range(
            header_row + 1,
            ws.max_row + 1,
        ):
            item = ws.cell(
                row=row,
                column=item_col,
            ).value

            qty = ws.cell(
                row=row,
                column=qty_col,
            ).value

            if item is None:
                continue

            item = str(item).strip()

            if item not in ITEM_MAP:
                continue

            ingredient = ITEM_MAP[item]

            try:
                qty = float(qty or 0)
            except (TypeError, ValueError):
                qty = 0.0

            result[ingredient] = (
                result.get(
                    ingredient,
                    0.0,
                )
                + qty
            )

        return result
    # ------------------------------------------------------
    # 입고 DB 반영
    # ------------------------------------------------------

    def apply_receipt(
        self,
        file_name,
        receipt,
    ):


        exists = self.db.fetchone(
            """
            SELECT id
            FROM receipt_history
            WHERE file_name = ?
            LIMIT 1
            """,
            (
                file_name,
            ),
        )

        if exists:

            print(
                "이미 등록된 입고파일입니다."
            )

            return False


        for ingredient, quantity in receipt.items():

            self.db.execute(
                """
                INSERT INTO receipt_history
                (
                    receipt_date,
                    file_name,
                    ingredient,
                    quantity
                )
                VALUES
                (
                    DATE('now'),
                    ?, ?, ?
                )
                """,
                (
                    file_name,
                    ingredient,
                    quantity,
                ),
            )

            self.db.execute(
                """
                UPDATE inventory_current
                SET
                    stock = stock + ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE ingredient = ?
                """,
                (
                    quantity,
                    ingredient,
                ),
            )

            today = datetime.today().strftime(
                "%Y-%m-%d"
            )

            self.db.execute(
                """
                UPDATE receipt_schedule
                SET
                    status='received'
                WHERE
                    ingredient=?
                    AND status='confirmed'
                    AND delivery_date<=?
                """,
                (
                    ingredient,
                    today,
                ),
            )

        return True

    # ------------------------------------------------------
    # 입 고  예 정  조 회
    # ------------------------------------------------------
    def get_expected_receipts(
        self,
        delivery_date=None,
    ):

        if delivery_date is None:

            from datetime import datetime

            delivery_date = datetime.today().strftime(
                "%Y-%m-%d"
            )

        rows = self.db.fetchall(
            """
            SELECT
                ingredient,
                quantity
            FROM receipt_schedule
            WHERE
                delivery_date=?
                AND status='confirmed'
            """,
            (
                delivery_date,
            ),
        )

        result = {}

        for row in rows:

            result[
                row["ingredient"]
            ] = row["quantity"]

        return result

    # ------------------------------------------------------
    # 입고 파일 처리
    # ------------------------------------------------------

    def import_file(
        self,
        file_path,
    ):

        receipt = self.import_excel(
            file_path,
        )

        success = self.apply_receipt(
            Path(file_path).name,
            receipt,
        )

        if not success:

            return None

        return receipt

# ==========================================================
# Global Engine
# ==========================================================

ENGINE = ReceiptEngine()


# ==========================================================
# Helper
# ==========================================================

def import_receipt(
    file_path,
):

    return ENGINE.import_file(
        file_path,
    )


# ==========================================================
# END OF FILE
# ==========================================================

