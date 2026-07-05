"""
==========================================================
HOMS Order Import Engine
==========================================================
교촌 발주 자동 입고
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from recipes.order_item_rules import (
    ORDER_ITEM_MAP,
)

from core.inventory_engine import InventoryEngine

from core.database_engine import DatabaseEngine

from core.receipt_engine import ReceiptEngine

ORDER_PATH = Path(
    "/home/hanul/koms_data/orders"
)


class OrderImportEngine:

    def __init__(
        self,
    ):
        self.inventory = InventoryEngine()

        self.db = DatabaseEngine()

        self.receipt = ReceiptEngine()

    def latest_files(self):

        files = sorted(
            ORDER_PATH.glob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )

        return files

    def raw_files(self):

        result = []

        for path in self.latest_files():

            wb = load_workbook(
                path,
                data_only=True,
            )

            ws = wb[wb.sheetnames[0]]

            title = str(ws["A1"].value)

            normalized = "".join(title.split())

            if "물류" not in normalized:
                continue

            result.append(path)

        return result
    def read_order_file(
        self,
        path,
    ):

        wb = load_workbook(
            path,
            data_only=True,
        )

        ws = wb[
            wb.sheetnames[0]
        ]

        result = {}

        for row in ws.iter_rows(
            min_row=5,
            values_only=True,
        ):

            if not row:
                continue

            if row[0] == "총  합 계 ":
                break

            name = row[1]

            qty = row[4]

            if name is None:
                continue

            if qty is None:
                continue

            result[str(name)] = int(qty)

        return result

    def convert_items(
        self,
        items,
    ):

        result = {}

        for name, qty in items.items():

            normalized = "".join(
                str(name).split()
            )

            for order_name, ingredient in ORDER_ITEM_MAP.items():

                key = "".join(
                    order_name.split()
                )

                if key in normalized:

                    result[ingredient] = (
                        result.get(
                            ingredient,
                            0,
                        )
                        + qty
                    )

                    break

        return result

    def import_orders(
        self,
    ):
        imported = {}

        for path in self.raw_files():

            if self.already_imported(
                path.name,
            ):
                continue

            items = self.read_order_file(
                path,
            )

            items = self.convert_items(
                items,
            )

            self.receipt.apply_receipt(
                path.name,
                items,
            )

            for ingredient, qty in items.items():

                imported[
                    ingredient
                ] = (
                    imported.get(
                        ingredient,
                        0,
                    )
                    + qty
                )

            self.mark_imported(
                path.name,
            )

        return imported

    def already_imported(
        self,
        file_name,
    ):

        row = self.db.fetchone(
            """
            SELECT id
            FROM import_history
            WHERE file_name=?
            """,
            (
                file_name,
            ),
        )

        return row is not None


    def mark_imported(
        self,
        file_name,
    ):

        self.db.execute(
            """
            INSERT INTO import_history
            (
                file_name
            )
            VALUES
            (
                ?
            )
            """,
            (
                file_name,
            ),
        )
