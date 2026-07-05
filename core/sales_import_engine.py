"""
==========================================================
HOMS Sales Import Engine
==========================================================
sales_2130 / sales_2355 / history Import
"""

import re

from core.text_utils import (
    normalize_menu,
)

from pathlib import Path

from openpyxl import load_workbook

from core.database_engine import DatabaseEngine

from core.path import KOMS_DATA_DIR

class SalesImportEngine:

    def __init__(self):
        self.db = DatabaseEngine()


    # ------------------------------------------------------
    # 매출 폴더
    # ------------------------------------------------------

    SALES_PATH = {

        "2130": KOMS_DATA_DIR
        / "sales_2130",

        "2355": KOMS_DATA_DIR
        / "sales_2355",

    }


    # ------------------------------------------------------
    # 파일 검색
    # ------------------------------------------------------

    def latest_files(
        self,
        source: str,
    ):

        base = self.SALES_PATH[
            source
        ]

        return sorted(

            base.rglob(
                "*.xlsx"
            ),

            key=lambda p:
                p.stat().st_mtime,

            reverse=False,

        )

    # ------------------------------------------------------
    # 판매 저장
    # ------------------------------------------------------
    def save_sale(
        self,
        sales_date: str,
        weekday: int,
        menu: str,
        qty: float,
        source: str,
    ):

        self.db.execute(
            """
            INSERT OR IGNORE INTO sales_history
            (
                sales_date,
                weekday,
                menu,
                qty,
                source
            )
            VALUES
            (
                ?, ?, ?, ?, ?
            )
            """,
            (
                sales_date,
                weekday,
                menu,
                qty,
                source,
            ),
        )

    # ------------------------------------------------------
    # 중복 확인
    # ------------------------------------------------------
    def has_sale(
        self,
        sales_date: str,
        menu: str,
        source: str,
    ) -> bool:

        row = self.db.fetchone(
            """
            SELECT COUNT(*) AS cnt
            FROM sales_history
            WHERE sales_date=?
              AND menu=?
              AND source=?
            """,
            (
                sales_date,
                menu,
                source,
            ),
        )

        return row["cnt"] > 0


    # ------------------------------------------------------
    # Import 여부
    # ------------------------------------------------------

    def already_imported(
        self,
        file_name: str,
    ) -> bool:

        row = self.db.fetchone(
            """
            SELECT id
            FROM import_history
            WHERE file_name=?
            """,
            (
                file_name,
            ),
        )

        return row is not None


    # ------------------------------------------------------
    # Import 기록
    # ------------------------------------------------------

    def mark_imported(
        self,
        file_name: str,
    ):

        self.db.execute(
            """
            INSERT OR IGNORE
            INTO import_history
            (
                file_name
            )
            VALUES
            (
                ?
            )
            """,
            (
                file_name,
            ),
        )

    # ------------------------------------------------------
    # 엑셀 한 개 Import
    # ------------------------------------------------------
    def import_excel(
        self,
        file_path,
        sales_date: str,
        source: str = "2355",
    ):

        from datetime import datetime

        weekday = datetime.strptime(
            sales_date,
            "%Y-%m-%d",
        ).weekday()

        wb = load_workbook(
            file_path,
            data_only=True,
        )

        ws = wb[wb.sheetnames[0]]

        saved = 0
        skipped = 0

        for row in ws.iter_rows(
            min_row=4,
            values_only=True,
        ):

            if not row:
                continue

            menu = normalize_menu(
                row[1],
            )

            qty = row[9]

            if not menu:
                continue

            if qty is None:
                continue

            try:
                qty = float(qty)
            except Exception:
                continue

            if qty <= 0:
                continue

            if self.has_sale(
                sales_date,
                menu,
                source,
            ):
                skipped += 1
                continue

            self.save_sale(
                sales_date,
                weekday,
                menu,
                qty,
                source,
            )

            saved += 1

        print()
        print("==============")
        print("IMPORT RESULT")
        print("==============")
        print("Saved   :", saved)
        print("Skipped :", skipped)
        if menu is None:
            return ""

        menu = str(menu)
        menu = "".join(menu.split())
        return menu

    # ------------------------------------------------------
    # 폴더 Import
    # ------------------------------------------------------

    def import_folder(
        self,
        source: str,
    ) -> int:

        imported = 0

        for path in self.latest_files(
            source,
        ):

            if self.already_imported(
                path.name,
            ):
                continue

            try:

                match = re.search(
                    r"(\d{8})",
                    path.stem,
                )

                if match is None:

                    print(
                        f"날짜를 찾을 수 없습니다: {path.name}"
                    )

                    continue

                date_text = match.group(1)

                sales_date = (
                    f"{date_text[:4]}-"
                    f"{date_text[4:6]}-"
                    f"{date_text[6:]}"
                )

                self.import_excel(
                    path,
                    sales_date,
                    source,
                )

                self.mark_imported(
                    path.name,
                )

                imported += 1

            except Exception as e:

                print(
                    f"[IMPORT ERROR] {path.name}"
                )

                print(e)

        return imported

