from pathlib import Path
from datetime import datetime
import sqlite3

from openpyxl import load_workbook

DB_PATH = Path.home() / "HOMS/database/HOMS.db"
SALES_DIR = Path.home() / "koms_data/sales_2355"

print("=" * 50)
print("HOMS SALES IMPORT")
print("=" * 50)

files = sorted(SALES_DIR.glob("*.xlsx"))

if not files:
    print("판매파일이 없습니다.")
    raise SystemExit

latest = files[-1]

print(f"파일 : {latest.name}")

wb = load_workbook(
    latest,
    data_only=True,
)

ws = wb[wb.sheetnames[0]]

sales_date = datetime.now().strftime("%Y-%m-%d")
weekday = datetime.now().weekday()

conn = sqlite3.connect(DB_PATH)
conn.row_factory = sqlite3.Row
cur = conn.cursor()

count = 0

for row in ws.iter_rows(min_row=4, values_only=True):

    if not row:
        continue

    menu = row[1]
    qty = row[9]

    if menu is None:
        continue

    menu = "".join(str(menu).split())

    if not menu:
        continue

    if qty is None:
        continue

    try:
        qty = float(qty)
    except Exception:
        continue

    if qty <= 0:
        continue

    cur.execute(
        """
        INSERT INTO sales_history
        (
            sales_date,
            weekday,
            menu,
            qty,
            source
        )
        VALUES
        (
            ?, ?, ?, ?, ?
        )
        """,
        (
            sales_date,
            weekday,
            menu,
            qty,
            "2355",
        ),
    )

    count += 1

conn.commit()
conn.close()

print()
print(f"{count}건 저장 완료")
